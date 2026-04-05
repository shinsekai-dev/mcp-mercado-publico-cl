"""Generador de Excel para cotizaciones de licitaciones."""

from pathlib import Path
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from domain.cotizacion.entities import Cotizacion
from infrastructure import brand


COLOR_HEADER_BG = brand.COLOR_PRIMARY
COLOR_HEADER_FG = brand.COLOR_WHITE
COLOR_SUBHEADER_BG = brand.COLOR_ACCENT
COLOR_ROW_ALT = brand.COLOR_BG_ALT
COLOR_TOTALES_BG = "E2EFDA"
COLOR_BORDER = brand.COLOR_BORDER


def _border(style="thin"):
    side = Side(style=style, color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _clp(value: float) -> str:
    """Formatea número como CLP con separador de miles."""
    return f"$ {value:,.0f}".replace(",", ".")


class CotizacionExcelGenerator:

    @staticmethod
    def generate(cotizacion: Cotizacion, output_path: Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"

        # Anchos de columna
        ws.column_dimensions["A"].width = 6    # N°
        ws.column_dimensions["B"].width = 45   # Descripción
        ws.column_dimensions["C"].width = 12   # Cantidad
        ws.column_dimensions["D"].width = 14   # Unidad
        ws.column_dimensions["E"].width = 18   # Precio Unit.
        ws.column_dimensions["F"].width = 18   # Total

        row = 1

        # ── Encabezado empresa ──────────────────────────────────────────────
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = cotizacion.proveedor.empresa.upper()
        ws[f"A{row}"].font = Font(name=brand.FONT_PRIMARY, bold=True, size=brand.FONT_SIZE_TITLE, color=COLOR_HEADER_FG)
        ws[f"A{row}"].fill = _fill(COLOR_HEADER_BG)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"RUT: {cotizacion.proveedor.rut}  |  {cotizacion.proveedor.giro or 'Servicios de tecnología'}"
        ws[f"A{row}"].font = Font(size=10, color=COLOR_HEADER_FG)
        ws[f"A{row}"].fill = _fill(COLOR_SUBHEADER_BG)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"{cotizacion.proveedor.direccion}  |  {cotizacion.proveedor.telefono}  |  {cotizacion.proveedor.email}"
        ws[f"A{row}"].font = Font(size=9, color=COLOR_HEADER_FG)
        ws[f"A{row}"].fill = _fill(COLOR_SUBHEADER_BG)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 2

        # ── Info licitación ─────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "COTIZACIÓN PARA LICITACIÓN PÚBLICA"
        ws[f"A{row}"].font = Font(bold=True, size=13, color=COLOR_HEADER_FG)
        ws[f"A{row}"].fill = _fill(COLOR_HEADER_BG)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        info_fields = [
            ("Código licitación:", cotizacion.codigo_licitacion),
            ("Nombre licitación:", cotizacion.nombre_licitacion),
            ("Organismo comprador:", cotizacion.organismo),
            ("Fecha cotización:", cotizacion.fecha.strftime("%d/%m/%Y")),
            ("Validez oferta:", f"{cotizacion.validez_oferta_dias} días corridos"),
            ("Moneda:", cotizacion.moneda),
        ]
        for label, value in info_fields:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True, size=10)
            ws.merge_cells(f"B{row}:F{row}")
            ws[f"B{row}"] = value
            ws[f"B{row}"].font = Font(size=10)
            row += 1

        row += 1

        # ── Tabla de ítems — header ─────────────────────────────────────────
        headers = ["N°", "Descripción", "Cantidad", "Unidad", "Precio Unit. Neto", "Total Neto"]
        col_letters = ["A", "B", "C", "D", "E", "F"]
        for col, (letter, header) in enumerate(zip(col_letters, headers), 1):
            cell = ws[f"{letter}{row}"]
            cell.value = header
            cell.font = Font(bold=True, size=10, color=COLOR_HEADER_FG)
            cell.fill = _fill(COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        # ── Filas de ítems ──────────────────────────────────────────────────
        items_start_row = row
        for i, item in enumerate(cotizacion.items):
            bg = COLOR_ROW_ALT if i % 2 == 1 else "FFFFFF"
            values = [
                item.correlativo,
                item.descripcion,
                item.cantidad,
                item.unidad_medida,
                item.precio_unitario_neto,
                item.precio_total_neto,
            ]
            alignments = ["center", "left", "center", "center", "right", "right"]
            for letter, value, align in zip(col_letters, values, alignments):
                cell = ws[f"{letter}{row}"]
                cell.value = value
                cell.font = Font(size=10)
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(letter == "B"))
                cell.border = _border()
                if letter in ("E", "F") and isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1

        # ── Totales ─────────────────────────────────────────────────────────
        totales = [
            ("Subtotal Neto:", cotizacion.subtotal_neto),
            ("IVA (19%):", cotizacion.iva),
            ("TOTAL:", cotizacion.total),
        ]
        for label, value in totales:
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True, size=10)
            ws[f"A{row}"].alignment = Alignment(horizontal="right")
            ws[f"A{row}"].fill = _fill(COLOR_TOTALES_BG)
            ws[f"A{row}"].border = _border()

            ws.merge_cells(f"E{row}:F{row}")
            ws[f"E{row}"] = value
            ws[f"E{row}"].font = Font(bold=(label == "TOTAL:"), size=10)
            ws[f"E{row}"].number_format = '#,##0'
            ws[f"E{row}"].alignment = Alignment(horizontal="right")
            ws[f"E{row}"].fill = _fill(COLOR_TOTALES_BG)
            ws[f"E{row}"].border = _border()
            row += 1

        row += 2

        # ── Observaciones ───────────────────────────────────────────────────
        if cotizacion.observaciones:
            ws[f"A{row}"] = "Observaciones:"
            ws[f"A{row}"].font = Font(bold=True, size=10)
            row += 1
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = cotizacion.observaciones
            ws[f"A{row}"].font = Font(size=10)
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 40
            row += 2

        # ── Firma ───────────────────────────────────────────────────────────
        ws.merge_cells(f"D{row}:F{row}")
        ws[f"D{row}"] = "_" * 35
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        row += 1
        ws.merge_cells(f"D{row}:F{row}")
        ws[f"D{row}"] = cotizacion.proveedor.representante_legal
        ws[f"D{row}"].font = Font(bold=True, size=10)
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        row += 1
        ws.merge_cells(f"D{row}:F{row}")
        ws[f"D{row}"] = "Representante Legal"
        ws[f"D{row}"].font = Font(size=9)
        ws[f"D{row}"].alignment = Alignment(horizontal="center")

        # Freeze header rows
        ws.freeze_panes = f"A{items_start_row}"

        wb.save(output_path)
        return output_path
